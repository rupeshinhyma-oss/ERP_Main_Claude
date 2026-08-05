"""
Generic Import / Export Helpers for Master Data.

Every master module supports CSV/Excel import and export. Rather than
re-implementing file parsing/serialization ten times, the row-level
parsing, validation-error accumulation, and workbook/CSV generation live
here once. A concrete module supplies:

    - a list of column headers (in the order they should appear),
    - a function that turns one raw row (dict[str, str]) into validated
      field kwargs (or raises ``BadRequestException``/``ValidationException``),
    - a function that turns one model instance into a row dict for export.

This keeps each master's ``service.py`` short: it calls
:func:`parse_import_file` and :func:`build_export_file`, and owns only the
per-row validation rules that are genuinely master-specific.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any, Callable
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import inspect as sa_inspect

from app.core.exceptions import AppException, BadRequestException, ConflictException


def model_to_dict(instance: Any) -> dict[str, Any]:
    """
    Serialize any SQLAlchemy ORM instance's plain columns into a JSON-safe dict.

    Generic (works for every model via introspection) so each service's
    duplicate-conflict handling doesn't need its own hand-written serializer
    just to show a "here's the existing record" comparison during import.
    """
    if instance is None:
        return {}
    mapper = sa_inspect(instance.__class__)
    result: dict[str, Any] = {}
    for column in mapper.columns:
        value = getattr(instance, column.key, None)
        if isinstance(value, (datetime, date)):
            value = value.isoformat()
        elif isinstance(value, Decimal):
            value = float(value)
        elif isinstance(value, UUID):
            value = str(value)
        elif isinstance(value, Enum):
            value = value.value
        result[column.key] = value
    return result


@dataclass
class ImportRowResult:
    """Outcome of validating a single imported row."""

    row_number: int  # 1-indexed, counting the header as row 1 (matches spreadsheet row numbers)
    success: bool
    field_values: dict[str, Any] | None = None
    error: str | None = None


@dataclass
class ImportSummary:
    """Aggregate result of an import run, returned to the caller and stored on the queue job.

    Nothing from the uploaded file is ever discarded: every row that fails
    (whether from a validation error or a duplicate conflict) is recorded in
    full, including its original raw values, so the caller can show, export,
    or let the user fix-and-retry every row that didn't make it in -- not
    just a row number and an error string.
    """

    total_rows: int = 0
    created: int = 0
    failed: int = 0
    duplicate_count: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)
    duplicates: list[dict[str, Any]] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        """Return a JSON-serializable summary dict."""
        return {
            "total_rows": self.total_rows,
            "created": self.created,
            "failed": self.failed,
            "duplicate_count": self.duplicate_count,
            "errors": self.errors,
            "duplicates": self.duplicates,
        }


def _read_csv_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    """Parse CSV bytes into a list of header->value dict rows."""
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_excel_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    """Parse the first worksheet of an .xlsx workbook into header->value dict rows."""
    workbook = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    rows: list[dict[str, str]] = []
    for raw_row in rows_iter:
        if raw_row is None or all(v is None for v in raw_row):
            continue
        row = {header[i]: ("" if raw_row[i] is None else str(raw_row[i])) for i in range(len(header)) if i < len(raw_row)}
        rows.append(row)
    return rows


def parse_rows_from_file(filename: str, raw_bytes: bytes) -> list[dict[str, str]]:
    """Parse an uploaded CSV or XLSX file into a list of header->value dict rows."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _read_csv_rows(raw_bytes)
    if lower.endswith(".xlsx"):
        return _read_excel_rows(raw_bytes)
    raise BadRequestException("Unsupported file type. Only .csv and .xlsx files are supported for import.")


async def run_import(
    rows: list[dict[str, str]],
    *,
    row_validator: Callable[[dict[str, str], int], dict[str, Any]],
    row_creator: Callable[[dict[str, Any]], Any],
) -> ImportSummary:
    """
    Validate and create every row, collecting a summary instead of failing the whole batch.

    Each row is validated and created independently: a bad row is recorded
    as a failure in the summary rather than aborting the import, so a
    typo in row 47 of a 500-row file doesn't waste the other 499 rows.

    No data loss: every row that doesn't get created -- whether it failed
    validation or collided with an existing record -- has its full original
    raw values preserved in the summary (``errors`` / ``duplicates``), not
    just a row number and a message. Rows that duplicate an existing record
    are additionally split into ``duplicates`` (rather than mixed into
    ``errors``) and, when the raising service attached the conflicting
    record (see :func:`model_to_dict`), the existing record's data is
    included too, so the caller can render a side-by-side comparison.

    Args:
        rows: Parsed rows (header -> raw string value).
        row_validator: Given a raw row and its 1-indexed row number, returns
            validated field kwargs ready for the repository, or raises
            :class:`~app.core.exceptions.BadRequestException` /
            :class:`~app.core.exceptions.ValidationException` /
            :class:`~app.core.exceptions.ConflictException` on invalid data.
        row_creator: Given validated field kwargs, creates and persists the record.
    """
    summary = ImportSummary(total_rows=len(rows))
    for index, raw_row in enumerate(rows, start=2):  # row 1 is the header
        try:
            field_values = row_validator(raw_row, index)
            await row_creator(field_values)
            summary.created += 1
        except ConflictException as exc:
            summary.failed += 1
            summary.duplicate_count += 1
            existing = exc.details.get("existing") if isinstance(exc.details, dict) else None
            summary.duplicates.append(
                {
                    "row": index,
                    "error": str(exc),
                    "row_data": dict(raw_row),
                    "existing": existing,
                }
            )
        except AppException as exc:  # noqa: BLE001 - one bad row must not abort the batch
            summary.failed += 1
            summary.errors.append({"row": index, "error": str(exc), "row_data": dict(raw_row)})
        except Exception as exc:  # noqa: BLE001 - catch-all so unexpected errors still don't abort the batch
            summary.failed += 1
            summary.errors.append({"row": index, "error": str(exc), "row_data": dict(raw_row)})
    return summary


def build_csv_export(headers: list[str], rows: list[dict[str, Any]]) -> bytes:
    """Build CSV file bytes from a list of header names and row dicts."""
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: ("" if row.get(h) is None else row.get(h)) for h in headers})
    return buffer.getvalue().encode("utf-8-sig")


def build_excel_export(headers: list[str], rows: list[dict[str, Any]], *, sheet_title: str = "Export") -> bytes:
    """Build .xlsx file bytes from a list of header names and row dicts."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]  # Excel sheet-title length limit
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(h) for h in headers])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
